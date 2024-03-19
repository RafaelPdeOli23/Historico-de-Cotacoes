from datetime import date
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
import os

#Cirando a pasta de saída
if os.path.exists('./saida'):
    pass
else:
    os.mkdir('./saida')

#Pegando o input do usuário
acao = input('Qual o código da Ação que você quer processar? ').upper()

#Coletando e tratando os dados do arquivo csv
try:
    if os.path.exists(f'./dados/{acao}.txt'):
        pass
    else:
        raise FileNotFoundError('Ação não encontrado')

    with open(f'./dados/{acao}.txt', 'r') as arquivo_cotacao:
        linhas = arquivo_cotacao.readlines()
        linhas = [linha.replace('\n', '').split(';') for linha in linhas]



    #Criando a planilha Excel auxiliar
    workbook = Workbook()
    planilha_ativa = workbook.active
    planilha_ativa.title = 'Dados'

    #Colocando um cabeçalho na tabela
    planilha_ativa.append(['DATA', 'COTAÇÃO', 'BANDA INFERIOR', 'BANDA SUPERIOR'])

    #Colocando os dados do arquivo csv para dentro da planilha 'Dados'
    indice = 2

    for linha in linhas:
        ano_mes_dia = linha[0].split(' ')[0]
        data = date(
            year=int(ano_mes_dia.split("-")[0]),
            month=int(ano_mes_dia.split("-")[1]),
            day=int(ano_mes_dia.split('-')[2]))
        cotacao = float(linha[1])

        #Atualiza as células da planilha ativa do Excel
        planilha_ativa[f'A{indice}'] = data
        planilha_ativa[f'B{indice}'] = cotacao

        indice += 1

    #Criando outra planilha dedicada para colocar um gráfico
    planilha_grafico = workbook.create_sheet('Gráficos')
    workbook.active = planilha_grafico

    #Colocando um título/cabeçalho no gráfico
    planilha_grafico.merge_cells('A1:T2')
    cabecalho = planilha_grafico['A1']
    cabecalho.font = Font(b=True, sz=18, color='FFFFFF')
    cabecalho.fill = PatternFill('solid', fgColor='ebd034')
    cabecalho.alignment = Alignment(vertical='center', horizontal='center')
    cabecalho.value = f'Histório de Cotações'

    #Criando o gráfico de cotações
    grafico = LineChart()
    grafico.width = 33.87
    grafico.height = 14.82
    grafico.title = F'{acao}'
    grafico.x_axis.tittle = 'Data da Cotação'
    grafico.y_axis.tittle = 'Valor da Cotação'

    referencia_cotacoes = Reference(planilha_ativa, min_col=2, min_row=2, max_col=2, max_row=indice)
    referencia_datas = Reference(planilha_ativa, min_col=1, min_row=2, max_col=1, max_row=indice)
    grafico.add_data(referencia_cotacoes)
    grafico.set_categories(referencia_datas)

    linha_cotacao = grafico.series[0]

    #Personalizando o gráfico
    linha_cotacao.graphicalProperties.line.width = 10
    linha_cotacao.graphicalProperties.line.solidfill = '3446eb'
    linha_cotacao.graphicalProperties.line.solidfill = 'ed1c2a'

    #Adicionando o gráfico à planilha
    planilha_grafico.add_chart(grafico, 'A3')

    #Efetivamente criando o arquivo Excel
    workbook.save('./saida/Planilha.xlsx')
    print('Planilha Excel criada!')

#Tratamento de um possível erro de digitação do papel da ação
except FileNotFoundError as erro:
    print(f'Ocorreu um erro: {erro}!')
    print('Verifique se escreveu o código da ação corretamente.')
